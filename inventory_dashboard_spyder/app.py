from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from functools import wraps
import sqlite3
import re
import calendar
from datetime import datetime, date
from io import BytesIO
import os
import openpyxl
from openpyxl.styles import Font
from werkzeug.security import generate_password_hash, check_password_hash

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader


app = Flask(__name__)

# Use an environment variable in production.
app.secret_key = os.environ.get(
    "FLASK_SECRET_KEY",
    "inventory-dashboard-secret"
)

DB_NAME = "inventory_boxes.db"

# Default login created the first time the app runs. Change this
# (or set DEFAULT_ADMIN_USERNAME / DEFAULT_ADMIN_PASSWORD env vars
# before first run) once you've logged in.
DEFAULT_ADMIN_USERNAME = os.environ.get("DEFAULT_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.environ.get("DEFAULT_ADMIN_PASSWORD", "admin123")

# Number of pallets pre-created on first run, purely for convenience.
# This is NOT a cap - pallets beyond this are created automatically
# the first time they're used (see add_product).
INITIAL_PALLETS = 10


# =========================================================
# DATABASE
# =========================================================

def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            box_weight REAL NOT NULL CHECK (box_weight > 0)
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS pallets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pallet_no INTEGER NOT NULL UNIQUE
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS stock (
            product_id INTEGER NOT NULL,
            pallet_id INTEGER NOT NULL,
            boxes REAL NOT NULL DEFAULT 0 CHECK (boxes >= 0),

            PRIMARY KEY (product_id, pallet_id),

            FOREIGN KEY (product_id)
                REFERENCES products(id)
                ON DELETE CASCADE,

            FOREIGN KEY (pallet_id)
                REFERENCES pallets(id)
                ON DELETE CASCADE
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            pallet_id INTEGER NOT NULL,
            movement_type TEXT NOT NULL
                CHECK (movement_type IN ('Inward', 'Outward')),
            boxes REAL NOT NULL CHECK (boxes > 0),
            created_at TEXT NOT NULL,
            transaction_date TEXT NOT NULL,

            FOREIGN KEY (product_id)
                REFERENCES products(id)
                ON DELETE CASCADE,

            FOREIGN KEY (pallet_id)
                REFERENCES pallets(id)
                ON DELETE CASCADE
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    """)

    # Batches track individual lots of stock with their own
    # manufacturing / expiry dates, so FIFO (nearest expiry first)
    # can be computed per product/pallet. "boxes" here is the
    # amount of this specific batch still remaining (it goes down
    # as Outward movements consume it).
    conn.execute("""
        CREATE TABLE IF NOT EXISTS batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            pallet_id INTEGER NOT NULL,
            customer_id INTEGER,
            mfg_date TEXT,
            expiry_date TEXT,
            boxes REAL NOT NULL DEFAULT 0 CHECK (boxes >= 0),
            transaction_date TEXT NOT NULL,
            created_at TEXT NOT NULL,

            FOREIGN KEY (product_id)
                REFERENCES products(id)
                ON DELETE CASCADE,

            FOREIGN KEY (pallet_id)
                REFERENCES pallets(id)
                ON DELETE CASCADE,

            FOREIGN KEY (customer_id)
                REFERENCES customers(id)
                ON DELETE SET NULL
        )
    """)

    # Migrate older databases that predate the customer / expiry
    # tracking columns on transactions. Safe to run every startup.
    existing_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(transactions)").fetchall()
    }

    for column, ddl in (
        ("customer_id", "ALTER TABLE transactions ADD COLUMN customer_id INTEGER"),
        ("mfg_date", "ALTER TABLE transactions ADD COLUMN mfg_date TEXT"),
        ("expiry_date", "ALTER TABLE transactions ADD COLUMN expiry_date TEXT"),
    ):
        if column not in existing_columns:
            try:
                conn.execute(ddl)
            except sqlite3.OperationalError:
                pass

    # Make sure P01-P10 exist.
    for i in range(1, INITIAL_PALLETS + 1):
        conn.execute("""
            INSERT OR IGNORE INTO pallets (pallet_no)
            VALUES (?)
        """, (i,))

    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL
        )
    """)

    # Seed a default login on first run only, so the app is never
    # left completely unprotected. Change the password after
    # logging in (or delete the row and re-run to reseed).
    existing_user = conn.execute(
        "SELECT id FROM users LIMIT 1"
    ).fetchone()

    if not existing_user:
        conn.execute("""
            INSERT INTO users (username, password_hash)
            VALUES (?, ?)
        """, (
            DEFAULT_ADMIN_USERNAME,
            generate_password_hash(DEFAULT_ADMIN_PASSWORD)
        ))

    conn.commit()
    conn.close()


# =========================================================
# HELPERS
# =========================================================

def close_quietly(conn):
    if conn is not None:
        try:
            conn.close()
        except Exception:
            pass


def format_date(date_string):
    try:
        return datetime.strptime(
            date_string,
            "%Y-%m-%d"
        ).strftime("%d-%b-%Y")
    except (TypeError, ValueError):
        return date_string


def parse_positive_float(value):
    try:
        number = float(str(value).strip())
        if number <= 0:
            return None
        return number
    except (TypeError, ValueError):
        return None


def parse_positive_int(value):
    try:
        number = int(str(value).strip())
        if number <= 0:
            return None
        return number
    except (TypeError, ValueError):
        return None


def valid_date(value):
    try:
        datetime.strptime(str(value).strip(), "%Y-%m-%d")
        return True
    except (TypeError, ValueError):
        return False


def period_range(period_type, period_value):
    """
    Turn a (period_type, period_value) pair into an inclusive
    (start_date, end_date) range of "YYYY-MM-DD" strings, or
    return None if the input isn't valid.

    period_type: "day" | "month" | "year"
    period_value:
        day   -> "YYYY-MM-DD"
        month -> "YYYY-MM"
        year  -> "YYYY"
    """
    try:
        if period_type == "day":
            d = datetime.strptime(period_value, "%Y-%m-%d").date()
            return d.isoformat(), d.isoformat()

        if period_type == "month":
            d = datetime.strptime(period_value, "%Y-%m").date()
            last_day = calendar.monthrange(d.year, d.month)[1]
            start = d.replace(day=1)
            end = d.replace(day=last_day)
            return start.isoformat(), end.isoformat()

        if period_type == "year":
            year = int(period_value)
            start = date(year, 1, 1)
            end = date(year, 12, 31)
            return start.isoformat(), end.isoformat()

    except (TypeError, ValueError):
        return None

    return None


def get_transaction_dates():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT DISTINCT transaction_date
            FROM transactions
            WHERE transaction_date IS NOT NULL
            ORDER BY transaction_date DESC
        """).fetchall()

        return [row["transaction_date"] for row in rows]
    finally:
        close_quietly(conn)


# =========================================================
# AUTH
# =========================================================

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.before_request
def require_login():
    # Anything not explicitly allowed needs a logged-in session.
    open_endpoints = {"login", "static"}

    if request.endpoint in open_endpoints:
        return None

    if not session.get("user_id"):
        return redirect(url_for("login", next=request.path))

    return None


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        conn = get_db()
        try:
            user = conn.execute("""
                SELECT id, username, password_hash
                FROM users
                WHERE username = ?
            """, (username,)).fetchone()
        finally:
            close_quietly(conn)

        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["username"] = user["username"]

            next_url = request.form.get("next") or url_for("dashboard")
            return redirect(next_url)

        flash("Invalid username or password.", "error")

    return render_template(
        "login.html",
        next=request.args.get("next", "")
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =========================================================
# DASHBOARD
# =========================================================

@app.route("/")
def dashboard():
    selected_date = request.args.get("date", "total")
    dates = get_transaction_dates()

    conn = get_db()

    try:
        all_products = conn.execute("""
            SELECT id, name, box_weight
            FROM products
            ORDER BY name COLLATE NOCASE
        """).fetchall()

        all_customers = conn.execute("""
            SELECT id, name
            FROM customers
            ORDER BY name COLLATE NOCASE
        """).fetchall()

        # -------------------------------------------------
        # TOTAL / CURRENT STOCK
        # -------------------------------------------------

        if selected_date == "total":

            products = conn.execute("""
                SELECT
                    p.id,
                    p.name,
                    p.box_weight,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN s.boxes > 0 THEN s.boxes
                                ELSE 0
                            END
                        ), 0
                    ) AS total_boxes,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN s.boxes > 0
                                THEN s.boxes * p.box_weight
                                ELSE 0
                            END
                        ), 0
                    ) AS total_weight,

                    COUNT(
                        CASE
                            WHEN s.boxes > 0 THEN 1
                        END
                    ) AS pallet_count

                FROM products p
                LEFT JOIN stock s
                    ON p.id = s.product_id

                GROUP BY p.id
                HAVING total_boxes > 0
                ORDER BY p.name COLLATE NOCASE
            """).fetchall()

            total_stock = conn.execute("""
                SELECT COALESCE(
                    SUM(s.boxes * p.box_weight), 0
                ) AS total
                FROM stock s
                JOIN products p
                    ON p.id = s.product_id
                WHERE s.boxes > 0
            """).fetchone()["total"]

            total_boxes = conn.execute("""
                SELECT COALESCE(SUM(boxes), 0) AS total
                FROM stock
                WHERE boxes > 0
            """).fetchone()["total"]

            total_inward = conn.execute("""
                SELECT COALESCE(
                    SUM(t.boxes * p.box_weight), 0
                ) AS total
                FROM transactions t
                JOIN products p
                    ON p.id = t.product_id
                WHERE t.movement_type = 'Inward'
            """).fetchone()["total"]

            total_outward = conn.execute("""
                SELECT COALESCE(
                    SUM(t.boxes * p.box_weight), 0
                ) AS total
                FROM transactions t
                JOIN products p
                    ON p.id = t.product_id
                WHERE t.movement_type = 'Outward'
            """).fetchone()["total"]

            display_date = None

            pallets_raw = conn.execute("""
                SELECT
                    pa.pallet_no,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN s.boxes > 0 THEN s.boxes
                                ELSE 0
                            END
                        ), 0
                    ) AS total_boxes,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN s.boxes > 0
                                THEN s.boxes * p.box_weight
                                ELSE 0
                            END
                        ), 0
                    ) AS total_weight

                FROM pallets pa
                LEFT JOIN stock s
                    ON pa.id = s.pallet_id
                LEFT JOIN products p
                    ON p.id = s.product_id

                GROUP BY pa.id
                ORDER BY pa.pallet_no
            """).fetchall()

            pallets = [dict(row) for row in pallets_raw]

        # -------------------------------------------------
        # DATE-WISE STOCK
        # -------------------------------------------------

        else:

            if not valid_date(selected_date):
                flash("Invalid date selected.", "error")
                return redirect(url_for("dashboard"))

            products_raw = conn.execute("""
                SELECT
                    p.id,
                    p.name,
                    p.box_weight,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN t.movement_type = 'Inward'
                                    THEN t.boxes
                                WHEN t.movement_type = 'Outward'
                                    THEN -t.boxes
                                ELSE 0
                            END
                        ), 0
                    ) AS total_boxes

                FROM products p
                LEFT JOIN transactions t
                    ON p.id = t.product_id
                    AND t.transaction_date <= ?

                GROUP BY p.id
                HAVING total_boxes > 0
                ORDER BY p.name COLLATE NOCASE
            """, (selected_date,)).fetchall()

            products = []

            for p in products_raw:
                boxes = p["total_boxes"]

                products.append({
                    "id": p["id"],
                    "name": p["name"],
                    "box_weight": p["box_weight"],
                    "total_boxes": boxes,
                    "total_weight": boxes * p["box_weight"],
                    "pallet_count": 0
                })

            total_stock = conn.execute("""
                SELECT COALESCE(
                    SUM(
                        CASE
                            WHEN t.movement_type = 'Inward'
                                THEN t.boxes * p.box_weight
                            WHEN t.movement_type = 'Outward'
                                THEN -t.boxes * p.box_weight
                            ELSE 0
                        END
                    ), 0
                ) AS total
                FROM transactions t
                JOIN products p
                    ON p.id = t.product_id
                WHERE t.transaction_date <= ?
            """, (selected_date,)).fetchone()["total"]

            total_boxes = conn.execute("""
                SELECT COALESCE(
                    SUM(
                        CASE
                            WHEN movement_type = 'Inward'
                                THEN boxes
                            WHEN movement_type = 'Outward'
                                THEN -boxes
                            ELSE 0
                        END
                    ), 0
                ) AS total
                FROM transactions
                WHERE transaction_date <= ?
            """, (selected_date,)).fetchone()["total"]

            total_inward = conn.execute("""
                SELECT COALESCE(
                    SUM(t.boxes * p.box_weight), 0
                ) AS total
                FROM transactions t
                JOIN products p
                    ON p.id = t.product_id
                WHERE t.movement_type = 'Inward'
                  AND t.transaction_date = ?
            """, (selected_date,)).fetchone()["total"]

            total_outward = conn.execute("""
                SELECT COALESCE(
                    SUM(t.boxes * p.box_weight), 0
                ) AS total
                FROM transactions t
                JOIN products p
                    ON p.id = t.product_id
                WHERE t.movement_type = 'Outward'
                  AND t.transaction_date = ?
            """, (selected_date,)).fetchone()["total"]

            display_date = format_date(selected_date)

            pallets_raw = conn.execute("""
                SELECT
                    pa.pallet_no,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN t.movement_type = 'Inward'
                                    THEN t.boxes
                                WHEN t.movement_type = 'Outward'
                                    THEN -t.boxes
                                ELSE 0
                            END
                        ), 0
                    ) AS total_boxes,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN t.movement_type = 'Inward'
                                    THEN t.boxes * p.box_weight
                                WHEN t.movement_type = 'Outward'
                                    THEN -t.boxes * p.box_weight
                                ELSE 0
                            END
                        ), 0
                    ) AS total_weight

                FROM pallets pa
                LEFT JOIN transactions t
                    ON pa.id = t.pallet_id
                    AND t.transaction_date <= ?
                LEFT JOIN products p
                    ON p.id = t.product_id

                GROUP BY pa.id
                ORDER BY pa.pallet_no
            """, (selected_date,)).fetchall()

            pallets = [dict(row) for row in pallets_raw]

        return render_template(
            "dashboard.html",
            products=products,
            all_products=all_products,
            all_customers=all_customers,
            pallets=pallets,
            total_stock=total_stock,
            total_boxes=total_boxes,
            total_inward=total_inward,
            total_outward=total_outward,
            dates=dates,
            selected_date=selected_date,
            display_date=display_date,
            today=date.today().strftime("%Y-%m-%d")
        )

    finally:
        close_quietly(conn)


def _read_excel_rows(file_storage):
    """
    Load the first sheet of an uploaded .xlsx file and return
    (header_list, data_rows). Returns (None, None) if the file
    can't be read as an Excel workbook.
    """
    try:
        workbook = openpyxl.load_workbook(
            file_storage,
            data_only=True,
            read_only=True
        )
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return None, None

    if not rows:
        return [], []

    header = [
        str(cell).strip().lower() if cell is not None else ""
        for cell in rows[0]
    ]

    return header, rows[1:]


def _find_column(header, *names):
    for name in names:
        if name in header:
            return header.index(name)
    return None


# =========================================================
# UPLOAD PRODUCTS (EXCEL)
# =========================================================
#
# Expected columns (case-insensitive): "name" and
# "box weight" (or "weight"). Existing products are updated
# in place by name; new names are inserted.
# =========================================================

@app.route("/upload_products", methods=["POST"])
def upload_products():

    file = request.files.get("products_file")

    if not file or not file.filename:
        flash("Please choose an Excel file to upload.", "error")
        return redirect(url_for("dashboard"))

    header, rows = _read_excel_rows(file)

    if header is None:
        flash(
            "Could not read that file. Please upload a valid .xlsx file.",
            "error"
        )
        return redirect(url_for("dashboard"))

    name_col = _find_column(header, "name", "product", "product name")
    weight_col = _find_column(
        header, "box weight", "box_weight", "weight"
    )

    if name_col is None or weight_col is None:
        flash(
            "The Excel file needs a 'Name' column and a "
            "'Box Weight' column.",
            "error"
        )
        return redirect(url_for("dashboard"))

    conn = get_db()
    added = 0
    skipped = 0

    try:
        for row in rows:

            if not row or name_col >= len(row):
                continue

            raw_name = row[name_col]

            if raw_name is None or not str(raw_name).strip():
                continue

            name = str(raw_name).strip()
            box_weight = parse_positive_float(
                row[weight_col] if weight_col < len(row) else None
            )

            if box_weight is None:
                skipped += 1
                continue

            try:
                conn.execute("""
                    INSERT INTO products (name, box_weight)
                    VALUES (?, ?)

                    ON CONFLICT(name)
                    DO UPDATE SET
                        box_weight = excluded.box_weight
                """, (name, box_weight))

                added += 1

            except sqlite3.Error:
                skipped += 1

        conn.commit()

        message = f"Imported {added} product(s) from Excel."
        if skipped:
            message += f" Skipped {skipped} invalid row(s)."

        flash(message, "success")
        return redirect(url_for("dashboard"))

    except sqlite3.Error:
        conn.rollback()
        flash("A database error occurred while importing products.", "error")
        return redirect(url_for("dashboard"))

    finally:
        close_quietly(conn)


# =========================================================
# UPLOAD CUSTOMERS (EXCEL)
# =========================================================
#
# Expected column (case-insensitive): "name" (or "customer",
# "customer name"). Duplicate names are ignored.
# =========================================================

@app.route("/upload_customers", methods=["POST"])
def upload_customers():

    file = request.files.get("customers_file")

    if not file or not file.filename:
        flash("Please choose an Excel file to upload.", "error")
        return redirect(url_for("dashboard"))

    header, rows = _read_excel_rows(file)

    if header is None:
        flash(
            "Could not read that file. Please upload a valid .xlsx file.",
            "error"
        )
        return redirect(url_for("dashboard"))

    name_col = _find_column(header, "name", "customer", "customer name")

    if name_col is None:
        flash(
            "The Excel file needs a 'Name' column.",
            "error"
        )
        return redirect(url_for("dashboard"))

    conn = get_db()
    added = 0

    try:
        for row in rows:

            if not row or name_col >= len(row):
                continue

            raw_name = row[name_col]

            if raw_name is None or not str(raw_name).strip():
                continue

            name = str(raw_name).strip()

            conn.execute("""
                INSERT OR IGNORE INTO customers (name)
                VALUES (?)
            """, (name,))

            added += 1

        conn.commit()

        flash(f"Imported {added} customer(s) from Excel.", "success")
        return redirect(url_for("dashboard"))

    except sqlite3.Error:
        conn.rollback()
        flash("A database error occurred while importing customers.", "error")
        return redirect(url_for("dashboard"))

    finally:
        close_quietly(conn)


# =========================================================
# API - PALLETS BY EXPIRY (FIFO)
# =========================================================
#
# Given a product, returns every pallet currently holding stock
# of that product, sorted so the pallet with the nearest expiry
# date comes first - i.e. the pallet that should be picked next
# under FIFO (First In, First Out).
# =========================================================

@app.route("/api/pallets_by_expiry")
def api_pallets_by_expiry():

    product_id = parse_positive_int(request.args.get("product_id", ""))

    if product_id is None:
        return jsonify({"error": "Invalid product."}), 400

    conn = get_db()

    try:
        rows = conn.execute("""
            SELECT
                pa.pallet_no,
                MIN(b.expiry_date) AS nearest_expiry,
                SUM(b.boxes) AS total_boxes

            FROM batches b
            JOIN pallets pa
                ON pa.id = b.pallet_id

            WHERE b.product_id = ?
              AND b.boxes > 0

            GROUP BY pa.id
            HAVING total_boxes > 0
            ORDER BY
                (nearest_expiry IS NULL),
                nearest_expiry ASC,
                pa.pallet_no ASC
        """, (product_id,)).fetchall()

        pallets = [
            {
                "pallet_no": row["pallet_no"],
                "nearest_expiry": row["nearest_expiry"],
                "total_boxes": row["total_boxes"]
            }
            for row in rows
        ]

        return jsonify({"pallets": pallets})

    finally:
        close_quietly(conn)


# =========================================================
# API - PRODUCTS FOR A CUSTOMER
# =========================================================
#
# Given a customer, returns the products they currently hold
# stock of (id, name, box_weight, total_boxes). Used by the Add
# Stock form so picking a customer can filter/prioritise the
# Product dropdown - e.g. Outward should only offer products
# that customer actually has.
# =========================================================

@app.route("/api/customer_products")
def api_customer_products():

    customer_id = parse_positive_int(request.args.get("customer_id", ""))

    if customer_id is None:
        return jsonify({"error": "Invalid customer."}), 400

    conn = get_db()

    try:
        rows = conn.execute("""
            SELECT
                p.id,
                p.name,
                p.box_weight,
                SUM(b.boxes) AS total_boxes

            FROM batches b
            JOIN products p
                ON p.id = b.product_id

            WHERE b.customer_id = ?
              AND b.boxes > 0

            GROUP BY p.id
            HAVING total_boxes > 0
            ORDER BY p.name COLLATE NOCASE
        """, (customer_id,)).fetchall()

        products = [
            {
                "id": row["id"],
                "name": row["name"],
                "box_weight": row["box_weight"],
                "total_boxes": row["total_boxes"]
            }
            for row in rows
        ]

        return jsonify({"products": products})

    finally:
        close_quietly(conn)


# =========================================================
# ADD PRODUCT / ADD STOCK
# =========================================================
#
# IMPORTANT FIX:
#
# CREATE PRODUCT:
#   Only name + box weight are required.
#
# ADD STOCK:
#   Product + boxes + pallet + date are required.
#
# The old version validated boxes/pallet/date BEFORE checking
# the mode, which prevented a new product from being created.
# =========================================================

@app.route("/add_product", methods=["POST"])
def add_product():

    product_id_text = request.form.get("product_id", "").strip()

    # -----------------------------------------------------
    # ADD STOCK / UPDATE EXISTING PRODUCT
    # -----------------------------------------------------
    #
    # pallet_no can contain MULTIPLE pallet numbers separated by
    # commas, spaces, and/or newlines (e.g. "1, 2, 5" or "1 2 5").
    # The same product/boxes/date/movement is applied to each
    # pallet listed. All pallets are updated in a single database
    # transaction - if any one of them fails, none of them are
    # applied.
    #
    # Inward movements create a new batch (with its own
    # manufacturing / expiry date) on each pallet listed.
    #
    # Outward movements consume existing batches FIFO - the
    # batch with the nearest expiry date on that pallet is drawn
    # down first.
    # -----------------------------------------------------

    product_id = parse_positive_int(product_id_text)
    boxes = parse_positive_float(
        request.form.get("boxes", "")
    )

    pallet_no_raw = request.form.get("pallet_no", "")

    # Split on commas, spaces, and newlines, then de-duplicate
    # while preserving the order the user typed them in.
    pallet_no_tokens = [
        token
        for token in re.split(r"[,\s]+", pallet_no_raw.strip())
        if token
    ]

    pallet_nos = []
    for token in pallet_no_tokens:
        parsed = parse_positive_int(token)

        if parsed is None:
            flash(
                f"'{token}' is not a valid pallet number.",
                "error"
            )
            return redirect(url_for("dashboard"))

        if parsed not in pallet_nos:
            pallet_nos.append(parsed)

    transaction_date = request.form.get(
        "transaction_date", ""
    ).strip()

    movement_type = request.form.get(
        "movement_type",
        "Inward"
    ).strip()

    customer_id = parse_positive_int(
        request.form.get("customer_id", "")
    )

    mfg_date = request.form.get("mfg_date", "").strip()
    expiry_date = request.form.get("expiry_date", "").strip()

    if product_id is None:
        flash("Please select a product.", "error")
        return redirect(url_for("dashboard"))

    if boxes is None:
        flash("Quantity must be greater than zero.", "error")
        return redirect(url_for("dashboard"))

    if not pallet_nos:
        flash(
            "Please enter at least one valid pallet number.",
            "error"
        )
        return redirect(url_for("dashboard"))

    if not valid_date(transaction_date):
        flash("Please enter a valid transaction date.", "error")
        return redirect(url_for("dashboard"))

    if movement_type not in ("Inward", "Outward"):
        flash("Invalid movement type.", "error")
        return redirect(url_for("dashboard"))

    if customer_id is None:
        flash("Please select a customer.", "error")
        return redirect(url_for("dashboard"))

    if movement_type == "Inward":

        if not valid_date(mfg_date):
            flash(
                "Please enter a valid manufacturing date.",
                "error"
            )
            return redirect(url_for("dashboard"))

        if not valid_date(expiry_date):
            flash("Please enter a valid expiry date.", "error")
            return redirect(url_for("dashboard"))

        if expiry_date < mfg_date:
            flash(
                "Expiry date can't be before the manufacturing date.",
                "error"
            )
            return redirect(url_for("dashboard"))

    conn = get_db()

    try:
        product = conn.execute("""
            SELECT id, name, box_weight
            FROM products
            WHERE id = ?
        """, (product_id,)).fetchone()

        if not product:
            flash("Product not found.", "error")
            return redirect(url_for("dashboard"))

        customer = conn.execute("""
            SELECT id, name
            FROM customers
            WHERE id = ?
        """, (customer_id,)).fetchone()

        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for("dashboard"))

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for pallet_no in pallet_nos:

            # Create the pallet on the fly if it doesn't exist yet -
            # pallet numbers are no longer capped.
            conn.execute("""
                INSERT OR IGNORE INTO pallets (pallet_no)
                VALUES (?)
            """, (pallet_no,))

            pallet = conn.execute("""
                SELECT id, pallet_no
                FROM pallets
                WHERE pallet_no = ?
            """, (pallet_no,)).fetchone()

            current = conn.execute("""
                SELECT boxes
                FROM stock
                WHERE product_id = ?
                  AND pallet_id = ?
            """, (
                product_id,
                pallet["id"]
            )).fetchone()

            current_boxes = (
                current["boxes"]
                if current
                else 0
            )

            txn_expiry_for_log = None

            if movement_type == "Inward":
                new_boxes = current_boxes + boxes

                # Record this lot as its own batch so FIFO can
                # find it later by expiry date.
                conn.execute("""
                    INSERT INTO batches (
                        product_id,
                        pallet_id,
                        customer_id,
                        mfg_date,
                        expiry_date,
                        boxes,
                        transaction_date,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    product_id,
                    pallet["id"],
                    customer_id,
                    mfg_date,
                    expiry_date,
                    boxes,
                    transaction_date,
                    created_at
                ))

                txn_expiry_for_log = expiry_date

            else:
                if boxes > current_boxes:
                    conn.rollback()
                    flash(
                        f"Cannot remove more stock than is currently "
                        f"on pallet P{pallet_no:02d}. No pallets in "
                        f"this request were updated.",
                        "error"
                    )
                    return redirect(url_for("dashboard"))

                new_boxes = current_boxes - boxes

                # FIFO: consume from the batch(es) on this pallet
                # with the nearest expiry date first.
                remaining_to_remove = boxes

                open_batches = conn.execute("""
                    SELECT id, boxes, expiry_date
                    FROM batches
                    WHERE product_id = ?
                      AND pallet_id = ?
                      AND boxes > 0
                    ORDER BY
                        (expiry_date IS NULL),
                        expiry_date ASC,
                        id ASC
                """, (
                    product_id,
                    pallet["id"]
                )).fetchall()

                earliest_expiry_consumed = None

                for batch in open_batches:

                    if remaining_to_remove <= 0:
                        break

                    take = min(batch["boxes"], remaining_to_remove)

                    conn.execute("""
                        UPDATE batches
                        SET boxes = boxes - ?
                        WHERE id = ?
                    """, (take, batch["id"]))

                    remaining_to_remove -= take

                    if earliest_expiry_consumed is None:
                        earliest_expiry_consumed = batch["expiry_date"]

                # If historical stock predates batch tracking, the
                # pallet's stock total can still be reduced even
                # when there isn't enough batch detail behind it.

                txn_expiry_for_log = earliest_expiry_consumed

            # Remove zero-stock rows instead of keeping unnecessary rows.
            if new_boxes == 0:
                conn.execute("""
                    DELETE FROM stock
                    WHERE product_id = ?
                      AND pallet_id = ?
                """, (
                    product_id,
                    pallet["id"]
                ))
            else:
                conn.execute("""
                    INSERT INTO stock (
                        product_id,
                        pallet_id,
                        boxes
                    )
                    VALUES (?, ?, ?)

                    ON CONFLICT(product_id, pallet_id)
                    DO UPDATE SET
                        boxes = excluded.boxes
                """, (
                    product_id,
                    pallet["id"],
                    new_boxes
                ))

            conn.execute("""
                INSERT INTO transactions (
                    product_id,
                    pallet_id,
                    movement_type,
                    boxes,
                    created_at,
                    transaction_date,
                    customer_id,
                    mfg_date,
                    expiry_date
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                product_id,
                pallet["id"],
                movement_type,
                boxes,
                created_at,
                transaction_date,
                customer_id,
                mfg_date if movement_type == "Inward" else None,
                txn_expiry_for_log
            ))

        conn.commit()

        if len(pallet_nos) == 1:
            flash(
                f"{movement_type} stock updated successfully.",
                "success"
            )
        else:
            pallet_list = ", ".join(
                f"P{no:02d}" for no in pallet_nos
            )
            flash(
                f"{movement_type} stock updated on {len(pallet_nos)} "
                f"pallets ({pallet_list}).",
                "success"
            )

        return redirect(url_for("dashboard"))

    except sqlite3.Error:
        conn.rollback()
        flash(
            "A database error occurred while updating stock.",
            "error"
        )
        return redirect(url_for("dashboard"))

    finally:
        close_quietly(conn)


# =========================================================
# PRODUCT PAGE
# =========================================================

@app.route("/product/<int:product_id>")
def product(product_id):

    # Optional ?date=YYYY-MM-DD - when present, show stock as it
    # stood at the end of that day plus that day's transactions,
    # instead of the live current stock / full history.
    selected_date = request.args.get("date", "").strip()

    if selected_date and not valid_date(selected_date):
        flash("Invalid date selected.", "error")
        return redirect(url_for("product", product_id=product_id))

    conn = get_db()

    try:
        product = conn.execute("""
            SELECT *
            FROM products
            WHERE id = ?
        """, (product_id,)).fetchone()

        if not product:
            return "Product not found", 404

        pallets = conn.execute("""
            SELECT
                pa.pallet_no,
                s.boxes AS boxes,
                (
                    SELECT MIN(b.expiry_date)
                    FROM batches b
                    WHERE b.product_id = s.product_id
                      AND b.pallet_id = s.pallet_id
                      AND b.boxes > 0
                ) AS nearest_expiry
            FROM pallets pa
            JOIN stock s
                ON pa.id = s.pallet_id
                AND s.product_id = ?
            WHERE s.boxes > 0
            ORDER BY
                (nearest_expiry IS NULL),
                nearest_expiry ASC,
                pa.pallet_no
        """, (product_id,)).fetchall()

        total_boxes = sum(
            p["boxes"] for p in pallets
        )

        total_weight = (
            total_boxes * product["box_weight"]
        )

        history = conn.execute("""
            SELECT
                t.movement_type,
                t.boxes,
                t.created_at,
                t.transaction_date,
                t.mfg_date,
                t.expiry_date,
                pa.pallet_no,
                c.name AS customer_name
            FROM transactions t
            JOIN pallets pa
                ON pa.id = t.pallet_id
            LEFT JOIN customers c
                ON c.id = t.customer_id
            WHERE t.product_id = ?
            ORDER BY
                t.transaction_date DESC,
                t.id DESC
        """, (product_id,)).fetchall()

        # -------------------------------------------------
        # DATE-SPECIFIC VIEW
        # -------------------------------------------------

        date_pallets = None
        date_transactions = None
        date_total_boxes = None
        date_total_weight = None
        display_date = None

        if selected_date:

            display_date = format_date(selected_date)

            date_pallets_raw = conn.execute("""
                SELECT
                    pa.pallet_no,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN t.movement_type = 'Inward'
                                    THEN t.boxes
                                WHEN t.movement_type = 'Outward'
                                    THEN -t.boxes
                                ELSE 0
                            END
                        ), 0
                    ) AS boxes

                FROM pallets pa
                LEFT JOIN transactions t
                    ON pa.id = t.pallet_id
                    AND t.product_id = ?
                    AND t.transaction_date <= ?

                GROUP BY pa.id
                HAVING boxes > 0
                ORDER BY pa.pallet_no
            """, (product_id, selected_date)).fetchall()

            date_pallets = [dict(row) for row in date_pallets_raw]

            date_total_boxes = sum(
                p["boxes"] for p in date_pallets
            )

            date_total_weight = (
                date_total_boxes * product["box_weight"]
            )

            date_transactions = conn.execute("""
                SELECT
                    t.movement_type,
                    t.boxes,
                    t.created_at,
                    pa.pallet_no
                FROM transactions t
                JOIN pallets pa
                    ON pa.id = t.pallet_id
                WHERE t.product_id = ?
                  AND t.transaction_date = ?
                ORDER BY t.id DESC
            """, (product_id, selected_date)).fetchall()

        all_customers = conn.execute("""
            SELECT id, name
            FROM customers
            ORDER BY name COLLATE NOCASE
        """).fetchall()

        return render_template(
            "product.html",
            product=product,
            pallets=pallets,
            selected_date=selected_date,
            display_date=display_date,
            date_pallets=date_pallets,
            date_transactions=date_transactions,
            date_total_boxes=date_total_boxes,
            date_total_weight=date_total_weight,
            total_boxes=total_boxes,
            total_weight=total_weight,
            history=history,
            all_customers=all_customers,
            today=date.today().strftime("%Y-%m-%d")
        )

    finally:
        close_quietly(conn)


# =========================================================
# PALLET PAGE
# =========================================================

@app.route("/pallet/<int:pallet_no>")
def pallet(pallet_no):

    if pallet_no < 1:
        return "Pallet not found", 404

    conn = get_db()

    try:
        pallet = conn.execute("""
            SELECT *
            FROM pallets
            WHERE pallet_no = ?
        """, (pallet_no,)).fetchone()

        if not pallet:
            return "Pallet not found", 404

        products = conn.execute("""
            SELECT
                p.id,
                p.name,
                p.box_weight,
                s.boxes,
                s.boxes * p.box_weight AS weight,
                (
                    SELECT MIN(b.expiry_date)
                    FROM batches b
                    WHERE b.product_id = p.id
                      AND b.pallet_id = s.pallet_id
                      AND b.boxes > 0
                ) AS nearest_expiry
            FROM stock s
            JOIN products p
                ON p.id = s.product_id
            WHERE s.pallet_id = ?
              AND s.boxes > 0
            ORDER BY p.name COLLATE NOCASE
        """, (pallet["id"],)).fetchall()

        total_boxes = sum(
            p["boxes"] for p in products
        )

        total_weight = sum(
            p["weight"] for p in products
        )

        return render_template(
            "pallet.html",
            pallet=pallet,
            products=products,
            total_boxes=total_boxes,
            total_weight=total_weight
        )

    finally:
        close_quietly(conn)

# =========================================================
# CUSTOMERS
# =========================================================

@app.route("/customers")
def customers():

    conn = get_db()

    try:
        customers = conn.execute("""
            SELECT
                c.id,
                c.name,

                COUNT(DISTINCT b.product_id) AS product_count,

                COALESCE(
                    SUM(
                        CASE WHEN b.boxes > 0 THEN b.boxes ELSE 0 END
                    ), 0
                ) AS total_boxes

            FROM customers c
            LEFT JOIN batches b
                ON b.customer_id = c.id
                AND b.boxes > 0

            GROUP BY c.id
            ORDER BY c.name COLLATE NOCASE
        """).fetchall()

        return render_template(
            "customers.html",
            customers=customers
        )

    finally:
        close_quietly(conn)


@app.route("/customer/<int:customer_id>")
def customer(customer_id):

    conn = get_db()

    try:
        customer = conn.execute("""
            SELECT id, name
            FROM customers
            WHERE id = ?
        """, (customer_id,)).fetchone()

        if not customer:
            return "Customer not found", 404

        # One row per open batch this customer has stock in -
        # i.e. the products they hold, broken down by pallet,
        # manufacturing date and expiry date.
        batches = conn.execute("""
            SELECT
                p.id AS product_id,
                p.name AS product_name,
                p.box_weight,
                pa.pallet_no,
                b.mfg_date,
                b.expiry_date,
                b.boxes

            FROM batches b
            JOIN products p
                ON p.id = b.product_id
            JOIN pallets pa
                ON pa.id = b.pallet_id

            WHERE b.customer_id = ?
              AND b.boxes > 0

            ORDER BY
                p.name COLLATE NOCASE,
                (b.expiry_date IS NULL),
                b.expiry_date ASC,
                pa.pallet_no
        """, (customer_id,)).fetchall()

        # Also group by product for a quick summary at the top -
        # "under the customer name, a list of the products they have".
        products_summary = {}
        for row in batches:
            key = row["product_id"]
            if key not in products_summary:
                products_summary[key] = {
                    "product_id": row["product_id"],
                    "product_name": row["product_name"],
                    "box_weight": row["box_weight"],
                    "total_boxes": 0
                }
            products_summary[key]["total_boxes"] += row["boxes"]

        products_summary = sorted(
            products_summary.values(),
            key=lambda item: item["product_name"].lower()
        )

        all_products = conn.execute("""
            SELECT id, name, box_weight
            FROM products
            ORDER BY name COLLATE NOCASE
        """).fetchall()

        return render_template(
            "customer.html",
            customer=customer,
            products_summary=products_summary,
            batches=batches,
            all_products=all_products,
            today=date.today().strftime("%Y-%m-%d")
        )

    finally:
        close_quietly(conn)


# =========================================================
# MANAGE DATA PAGE
# =========================================================

@app.route("/manage-data")
def manage_data():

    conn = get_db()

    try:

        all_products = conn.execute("""
            SELECT id, name, box_weight
            FROM products
            ORDER BY name COLLATE NOCASE
        """).fetchall()

        return render_template(
            "manage_data.html",
            all_products=all_products
        )

    finally:

        close_quietly(conn)
# =========================================================
# CLEAR PALLET
# =========================================================
#
# Unlike the old version, this records the removed stock as
# Outward transactions before clearing the current stock.
# This keeps current stock and transaction history consistent.
# =========================================================

@app.route("/clear_pallet", methods=["POST"])
def clear_pallet():

    pallet_no = parse_positive_int(
        request.form.get("pallet_no", "")
    )

    if pallet_no is None:
        flash("Invalid pallet number.", "error")
        return redirect(url_for("dashboard"))

    conn = get_db()

    try:
        pallet = conn.execute("""
            SELECT *
            FROM pallets
            WHERE pallet_no = ?
        """, (pallet_no,)).fetchone()

        if not pallet:
            flash("Pallet not found.", "error")
            return redirect(url_for("dashboard"))

        rows = conn.execute("""
            SELECT product_id, boxes
            FROM stock
            WHERE pallet_id = ?
              AND boxes > 0
        """, (pallet["id"],)).fetchall()

        if not rows:
            flash(
                f"P{pallet_no:02d} is already empty.",
                "error"
            )
            return redirect(url_for("dashboard"))

        transaction_date = date.today().strftime("%Y-%m-%d")
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Record everything being removed.
        for row in rows:
            conn.execute("""
                INSERT INTO transactions (
                    product_id,
                    pallet_id,
                    movement_type,
                    boxes,
                    created_at,
                    transaction_date
                )
                VALUES (?, ?, 'Outward', ?, ?, ?)
            """, (
                row["product_id"],
                pallet["id"],
                row["boxes"],
                created_at,
                transaction_date
            ))

        conn.execute("""
            DELETE FROM stock
            WHERE pallet_id = ?
        """, (pallet["id"],))

        conn.commit()

        flash(
            f"P{pallet_no:02d} has been cleared and the removed stock was recorded.",
            "success"
        )

        return redirect(url_for("dashboard"))

    except sqlite3.Error:
        conn.rollback()
        flash(
            "A database error occurred while clearing the pallet.",
            "error"
        )
        return redirect(url_for("dashboard"))

    finally:
        close_quietly(conn)


# =========================================================
# DELETE PRODUCT
# =========================================================

@app.route("/delete_product/<int:product_id>", methods=["POST"])
def delete_product(product_id):

    conn = get_db()

    try:
        product = conn.execute("""
            SELECT id, name
            FROM products
            WHERE id = ?
        """, (product_id,)).fetchone()

        if not product:
            flash("Product not found.", "error")
            return redirect(url_for("dashboard"))

        # Do not silently delete a product that still has stock.
        stock = conn.execute("""
            SELECT COALESCE(SUM(boxes), 0) AS total
            FROM stock
            WHERE product_id = ?
        """, (product_id,)).fetchone()

        if stock["total"] > 0:
            flash(
                "Cannot delete a product that still has stock. Remove the stock first.",
                "error"
            )
            return redirect(url_for("dashboard"))

        conn.execute("""
            DELETE FROM products
            WHERE id = ?
        """, (product_id,))

        conn.commit()

        flash(
            f"Product '{product['name']}' deleted successfully.",
            "success"
        )

        return redirect(url_for("dashboard"))

    except sqlite3.Error:
        conn.rollback()
        flash(
            "A database error occurred while deleting the product.",
            "error"
        )
        return redirect(url_for("dashboard"))

    finally:
        close_quietly(conn)


# =========================================================
# CLEAR ALL DATA
# =========================================================

@app.route("/clear_all_data", methods=["POST"])
def clear_all_data():

    conn = get_db()

    try:
        conn.execute("DELETE FROM transactions")
        conn.execute("DELETE FROM stock")
        conn.execute("DELETE FROM products")

        conn.commit()

        flash(
            "All inventory data has been deleted.",
            "success"
        )

        return redirect(url_for("dashboard"))

    except sqlite3.Error:
        conn.rollback()
        flash(
            "A database error occurred while clearing the data.",
            "error"
        )
        return redirect(url_for("dashboard"))

    finally:
        close_quietly(conn)


# =========================================================
# API - GOODS FOR A GIVEN DATE / MOVEMENT TYPE
# =========================================================
# Used by the Export (GRN) page to auto-fill the Goods Details
# table when the user picks a date and Inward/Outward. Returns
# one row per product, with the boxes summed across all pallets
# for that date + movement type.

@app.route("/api/grn_items")
def api_grn_items():

    item_date = request.args.get("date", "").strip()
    movement_type = request.args.get("movement_type", "").strip()

    if not valid_date(item_date):
        return jsonify({
            "error": "Please provide a valid date."
        }), 400

    if movement_type not in ("Inward", "Outward"):
        return jsonify({
            "error": "movement_type must be 'Inward' or 'Outward'."
        }), 400

    conn = get_db()

    try:
        rows = conn.execute("""
            SELECT
                p.id AS product_id,
                p.name AS name,
                p.box_weight AS weight,
                SUM(t.boxes) AS quantity

            FROM transactions t
            JOIN products p
                ON p.id = t.product_id

            WHERE t.transaction_date = ?
              AND t.movement_type = ?

            GROUP BY p.id
            ORDER BY p.name COLLATE NOCASE
        """, (item_date, movement_type)).fetchall()

        items = [
            {
                "product_id": row["product_id"],
                "name": row["name"],
                "weight": row["weight"],
                "quantity": row["quantity"]
            }
            for row in rows
        ]

        return jsonify({"items": items})

    finally:
        close_quietly(conn)


# =========================================================
# EXPORT PAGE
# =========================================================

@app.route("/export")
def export_page():

    conn = get_db()

    try:
        products = conn.execute("""
            SELECT
                id,
                name,
                box_weight
            FROM products
            ORDER BY name COLLATE NOCASE
        """).fetchall()

        return render_template(
            "Export.html",
            products=products,
            today=date.today().strftime("%Y-%m-%d")
        )

    finally:
        close_quietly(conn)


# =========================================================
# EXCEL STOCK EXPORT (day / month / year)
# =========================================================

@app.route("/export_excel")
def export_excel_page():
    return render_template(
        "export_excel.html",
        today=date.today().strftime("%Y-%m-%d"),
        this_month=date.today().strftime("%Y-%m"),
        this_year=date.today().strftime("%Y")
    )


@app.route("/export_excel/download")
def export_excel_download():

    period_type = request.args.get("period_type", "").strip()
    period_value = request.args.get("period_value", "").strip()

    if period_type not in ("day", "month", "year"):
        flash("Please choose a valid period.", "error")
        return redirect(url_for("export_excel_page"))

    date_range = period_range(period_type, period_value)

    if date_range is None:
        flash("Please choose a valid date.", "error")
        return redirect(url_for("export_excel_page"))

    start_date, end_date = date_range

    conn = get_db()

    try:
        # Stock received within the chosen period that is still on
        # hand (boxes > 0), one row per batch/pallet - this is the
        # "total stock for a day / month / year" the customer holds.
        rows = conn.execute("""
            SELECT
                COALESCE(c.name, 'Unassigned') AS customer_name,
                p.name AS product_name,
                b.mfg_date,
                b.expiry_date,
                b.boxes AS quantity,
                pa.pallet_no

            FROM batches b
            JOIN products p
                ON p.id = b.product_id
            JOIN pallets pa
                ON pa.id = b.pallet_id
            LEFT JOIN customers c
                ON c.id = b.customer_id

            WHERE b.boxes > 0
              AND b.transaction_date BETWEEN ? AND ?

            ORDER BY
                customer_name COLLATE NOCASE,
                p.name COLLATE NOCASE,
                pa.pallet_no
        """, (start_date, end_date)).fetchall()

    finally:
        close_quietly(conn)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Stock"

    headers = [
        "Customer Name",
        "Product Name",
        "Manufacturing Date",
        "Expiry Date",
        "Quantity",
        "Pallet Number",
    ]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([
            row["customer_name"],
            row["product_name"],
            format_date(row["mfg_date"]),
            format_date(row["expiry_date"]),
            row["quantity"],
            f"P{row['pallet_no']:02d}",
        ])

    for column_cells in sheet.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"stock_export_{period_type}_{period_value}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================================
# PDF HELPERS
# =========================================================

def draw_wrapped_text(pdf, text, x, y, max_width, line_height=9, font="Helvetica", size=7):
    """
    Draw simple wrapped text and return the new y position.
    """
    text = "" if text is None else str(text)

    pdf.setFont(font, size)

    words = text.split()
    if not words:
        return y

    lines = []
    current = ""

    for word in words:
        test = word if not current else current + " " + word

        if pdf.stringWidth(test, font, size) <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word

    if current:
        lines.append(current)

    for line in lines:
        pdf.drawString(x, y, line)
        y -= line_height

    return y


def draw_grn_header(pdf, width, height, voucher_no, grn_date, ref,
                    vehicle_no, credit_days, driver_name, product_temp,
                    out_time, driver_contact, in_time):

    pdf.setLineWidth(1)

    pdf.rect(
        20,
        25,
        width - 40,
        height - 50
    )

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(
        width / 2,
        height - 42,
        "GOODS RECEIVED NOTE"
    )

    top_y = height - 55

    pdf.rect(
        30,
        top_y - 90,
        width - 60,
        90
    )

    pdf.setFont("Helvetica-Bold", 8.5)

    pdf.drawString(
        38,
        top_y - 13,
        "BOBBA AVIATION CARGO AND GROUND HANDLING SERVICES PRIVATE"
    )

    pdf.drawString(
        38,
        top_y - 23,
        "LIMITED"
    )

    pdf.setFont("Helvetica", 7.5)

    pdf.drawString(
        38,
        top_y - 34,
        "SY.No.22, MVIT College Road Sonnappanahalli Village, Jala Hobli"
    )

    pdf.drawString(
        38,
        top_y - 44,
        "Bangalore - 562157"
    )

    pdf.drawString(
        38,
        top_y - 54,
        "Karnataka (29) - India"
    )

    pdf.drawString(
        38,
        top_y - 64,
        "CIN No.: U62200KA2007PTC044812"
    )

    pdf.drawString(
        38,
        top_y - 74,
        "PAN No.: AADCB3096R"
    )

    pdf.drawString(
        38,
        top_y - 84,
        "GST No.: 29AADCB3096R1ZX"
    )

    # -----------------------------------------------------
    # LOGO
    # -----------------------------------------------------
    # Look for the logo case-insensitively and allow the file
    # extension to vary. This prevents the PDF from silently
    # losing the logo when Windows saved it as .PNG or when the
    # filename contains an extra extension.
    logo_dir = os.path.join(app.root_path, "static")
    logo_candidates = []

    if os.path.isdir(logo_dir):
        for filename in os.listdir(logo_dir):
            stem, ext = os.path.splitext(filename)
            if stem.lower() == "bobba_logo" and ext.lower() in (".png", ".jpg", ".jpeg"):
                logo_candidates.append(os.path.join(logo_dir, filename))

    if logo_candidates:
        try:
            logo_path = logo_candidates[0]
            logo = ImageReader(logo_path)
            logo_w = 145
            logo_h = 52
            logo_x = width - 30 - logo_w
            logo_y = top_y - 66

            pdf.drawImage(
                logo,
                logo_x,
                logo_y,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                anchor="sw",
                mask="auto"
            )
        except Exception as exc:
            print(f"Warning: could not load Bobba logo: {exc}")

    receiver_top = top_y - 90
    receiver_height = 120
    right_x = 400

    pdf.rect(
        30,
        receiver_top - receiver_height,
        width - 60,
        receiver_height
    )

    pdf.line(
        right_x,
        receiver_top,
        right_x,
        receiver_top - receiver_height
    )

    pdf.setFont("Helvetica-Bold", 8.5)

    pdf.drawString(
        38,
        receiver_top - 15,
        "FARMEDIBLE PRODUCTS"
    )

    pdf.drawString(
        38,
        receiver_top - 25,
        "PRIVATE LIMITED"
    )

    pdf.setFont("Helvetica", 7.5)

    pdf.drawString(
        38,
        receiver_top - 37,
        "Plot No.64, Hotagalli Industrial Area,"
    )

    pdf.drawString(
        38,
        receiver_top - 47,
        "Mysuru - 570018"
    )

    pdf.drawString(
        38,
        receiver_top - 57,
        "Karnataka (29) - India"
    )

    pdf.drawString(
        38,
        receiver_top - 67,
        "9845537618"
    )

    pdf.drawString(
        38,
        receiver_top - 77,
        "PAN No.: AADCF8552C"
    )

    pdf.drawString(
        38,
        receiver_top - 87,
        "GST No.: 29AADCF8552C1ZL"
    )

    labels = [
        ("Voucher No.", voucher_no),
        ("Date", grn_date),
        ("Ref", ref),
        ("Vehicle No.", vehicle_no),
        ("Credit Days", credit_days),
        ("Driver Name", driver_name),
        ("Product Temp", product_temp),
        ("Out Time", out_time),
        ("Driver Contact No.", driver_contact),
        ("In Time", in_time),
    ]

    y = receiver_top - 12

    for label, value in labels:
        pdf.setFont("Helvetica-Bold", 7.5)
        pdf.drawString(right_x + 5, y, label)

        pdf.setFont("Helvetica", 7.5)
        pdf.drawString(
            right_x + 80,
            y,
            str(value)
        )

        y -= 11

    return receiver_top - receiver_height


# =========================================================
# GENERATE GRN PDF
# =========================================================

@app.route("/generate_grn", methods=["POST"])
def generate_grn():

    voucher_no = request.form.get("voucher_no", "").strip()
    grn_date = request.form.get("grn_date", "").strip()
    ref = request.form.get("ref", "").strip()
    vehicle_no = request.form.get("vehicle_no", "").strip()
    credit_days = request.form.get("credit_days", "").strip()
    driver_name = request.form.get("driver_name", "").strip()
    product_temp = request.form.get("product_temp", "").strip()
    out_time = request.form.get("out_time", "").strip()
    driver_contact = request.form.get("driver_contact", "").strip()
    in_time = request.form.get("in_time", "").strip()
    packing_slip = request.form.get("packing_slip", "").strip()
    number_boxes = request.form.get("number_boxes", "").strip()

    product_ids = request.form.getlist("product_id[]")
    quantities = request.form.getlist("quantity[]")
    order_units = request.form.getlist("order_unit[]")
    batches = request.form.getlist("batch[]")
    mfg_dates = request.form.getlist("mfg_date[]")
    expiry_dates = request.form.getlist("expiry[]")

    if not voucher_no:
        flash("Voucher number is required.", "error")
        return redirect(url_for("export_page"))

    if not grn_date or not valid_date(grn_date):
        flash("Please enter a valid GRN date.", "error")
        return redirect(url_for("export_page"))

    if not product_ids:
        flash("Please add at least one product to the GRN.", "error")
        return redirect(url_for("export_page"))

    conn = get_db()

    try:
        items = []

        for i, raw_product_id in enumerate(product_ids):

            if not raw_product_id.strip():
                continue

            if i >= len(quantities):
                flash(
                    f"Quantity is missing for product row {i + 1}.",
                    "error"
                )
                return redirect(url_for("export_page"))

            product_id = parse_positive_int(raw_product_id)
            quantity = parse_positive_float(quantities[i])

            if product_id is None:
                flash(
                    f"Invalid product in row {i + 1}.",
                    "error"
                )
                return redirect(url_for("export_page"))

            if quantity is None:
                flash(
                    f"Quantity must be greater than zero in row {i + 1}.",
                    "error"
                )
                return redirect(url_for("export_page"))

            product = conn.execute("""
                SELECT id, name, box_weight
                FROM products
                WHERE id = ?
            """, (product_id,)).fetchone()

            if not product:
                flash(
                    f"Product in row {i + 1} was not found.",
                    "error"
                )
                return redirect(url_for("export_page"))

            weight_per_box = product["box_weight"]
            net_weight = quantity * weight_per_box

            items.append({
                "description": product["name"],
                "weight": weight_per_box,
                "batch": (
                    batches[i]
                    if i < len(batches)
                    else ""
                ),
                "mfg": (
                    mfg_dates[i]
                    if i < len(mfg_dates)
                    else ""
                ),
                "expiry": (
                    expiry_dates[i]
                    if i < len(expiry_dates)
                    else ""
                ),
                "qty": quantity,
                "unit": (
                    order_units[i]
                    if i < len(order_units) and order_units[i]
                    else "Box"
                ),
                "net_weight": net_weight
            })

        if not items:
            flash("Please add at least one valid product.", "error")
            return redirect(url_for("export_page"))

    finally:
        close_quietly(conn)

    total_qty = sum(
        item["qty"] for item in items
    )

    total_weight = sum(
        item["net_weight"] for item in items
    )

    # =====================================================
    # CREATE PDF
    # =====================================================

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=A4
    )

    width, height = A4

    table_left = 30
    table_right = width - 30

    # A4 has limited space. Use multiple pages if necessary.
    rows_per_page = 9

    for page_start in range(
        0,
        len(items),
        rows_per_page
    ):

        page_items = items[
            page_start:page_start + rows_per_page
        ]

        table_top = draw_grn_header(
            pdf,
            width,
            height,
            voucher_no,
            grn_date,
            ref,
            vehicle_no,
            credit_days,
            driver_name,
            product_temp,
            out_time,
            driver_contact,
            in_time
        )

        # -------------------------------------------------
        # GOODS TABLE
        # -------------------------------------------------

        header_height = 25
        row_height = 24
        table_height = (
            header_height +
            row_height * rows_per_page +
            22
        )

        table_bottom = table_top - table_height

        pdf.rect(
            table_left,
            table_bottom,
            table_right - table_left,
            table_height
        )

        # All columns MUST stay inside the A4 page.
        # A4 width is ~595 pt; with 30 pt margins the usable
        # width is ~535 pt.
        columns = [
            30,   # S.No
            58,   # Description
            230,  # Weight Per
            285,  # Batch
            335,  # MFG Date
            385,  # Expiry
            435,  # Qty
            475,  # Ord. Unit
            515,  # Net Weight
            565   # right edge
        ]

        for x in columns[1:]:
            pdf.line(
                x,
                table_bottom,
                x,
                table_top
            )

        pdf.line(
            table_left,
            table_top - header_height,
            table_right,
            table_top - header_height
        )

        headers = [
            "S.No",
            "Description of Goods",
            "Weight Per",
            "Batch",
            "MFG Date",
            "Expiry",
            "Qty",
            "Ord. Unit",
            "Net Weight"
        ]

        header_x = [
            34,
            62,
            234,
            289,
            339,
            389,
            439,
            479,
            519
        ]

        pdf.setFont(
            "Helvetica-Bold",
            6.5
        )

        for x, text in zip(
            header_x,
            headers
        ):
            pdf.drawString(
                x,
                table_top - 10,
                text
            )

        # -------------------------------------------------
        # TABLE ROWS
        # -------------------------------------------------

        y = (
            table_top -
            header_height -
            16
        )

        pdf.setFont(
            "Helvetica",
            6.8
        )

        for local_index, item in enumerate(page_items):

            global_index = page_start + local_index

            pdf.drawString(
                35,
                y,
                str(global_index + 1)
            )

            description = str(
                item["description"]
            )

            # Wrap description inside its column.
            desc_words = description.split()
            desc_line = ""

            if desc_words:
                for word in desc_words:
                    candidate = (
                        word
                        if not desc_line
                        else desc_line + " " + word
                    )

                    if pdf.stringWidth(
                        candidate,
                        "Helvetica",
                        6.8
                    ) <= 160:
                        desc_line = candidate
                    else:
                        break

            pdf.drawString(
                62,
                y,
                desc_line[:30]
            )

            pdf.drawRightString(
                280,
                y,
                f"{item['weight']:.2f}"
            )

            pdf.drawString(
                289,
                y,
                str(item["batch"])[:8]
            )

            pdf.drawString(
                339,
                y,
                str(item["mfg"])[:8]
            )

            pdf.drawString(
                389,
                y,
                str(item["expiry"])[:8]
            )

            pdf.drawRightString(
                471,
                y,
                f"{item['qty']:.2f}"
            )

            pdf.drawString(
                479,
                y,
                str(item["unit"])[:6]
            )

            pdf.drawRightString(
                561,
                y,
                f"{item['net_weight']:.2f}"
            )

            y -= row_height

        # -------------------------------------------------
        # TOTAL ONLY ON LAST PAGE
        # -------------------------------------------------

        is_last_page = (
            page_start + rows_per_page >= len(items)
        )

        if is_last_page:

            pdf.line(
                table_left,
                table_bottom + 22,
                table_right,
                table_bottom + 22
            )

            pdf.setFont(
                "Helvetica-Bold",
                8
            )

            pdf.drawRightString(
                425,
                table_bottom + 7,
                "Total"
            )

            pdf.drawRightString(
                471,
                table_bottom + 7,
                f"{total_qty:.2f}"
            )

            pdf.drawRightString(
                561,
                table_bottom + 7,
                f"{total_weight:.2f}"
            )

            # ---------------------------------------------
            # PACKING SLIP
            # ---------------------------------------------

            slip_y = table_bottom - 30

            pdf.rect(
                30,
                slip_y,
                width - 60,
                30
            )

            pdf.setFont(
                "Helvetica",
                8
            )

            pdf.drawString(
                35,
                slip_y + 18,
                "Packing Slip No.:"
            )

            pdf.drawString(
                150,
                slip_y + 18,
                packing_slip
            )

            pdf.drawString(
                330,
                slip_y + 18,
                "No. of Boxes:"
            )

            pdf.drawString(
                420,
                slip_y + 18,
                number_boxes
            )

            # ---------------------------------------------
            # TERMS AND CONDITIONS
            # ---------------------------------------------

            terms_y = slip_y - 110

            pdf.rect(
                30,
                terms_y,
                width - 60,
                105
            )

            pdf.setFont(
                "Helvetica-Bold",
                8.5
            )

            pdf.drawString(
                38,
                terms_y + 88,
                "Terms and Conditions"
            )

            pdf.setFont(
                "Helvetica",
                8
            )

            pdf.drawString(
                38,
                terms_y + 73,
                "Quality Certificate from Manufacture Enclosed: YES/NO"
            )

            pdf.drawString(
                38,
                terms_y + 56,
                "Enclosed:"
            )

            pdf.drawString(
                55,
                terms_y + 43,
                "1.Supplier Bill"
            )

            pdf.drawString(
                38,
                terms_y + 12,
                "Customer Sign:"
            )

            pdf.drawString(
                210,
                terms_y + 12,
                "Received By:"
            )

            pdf.drawString(
                365,
                terms_y + 12,
                "Checked By:"
            )

            pdf.line(
                30,
                terms_y,
                width - 30,
                terms_y
            )

            pdf.setFont(
                "Helvetica-Bold",
                8.5
            )

            pdf.drawCentredString(
                width / 2,
                terms_y - 15,
                "For BOBBA AVIATION CARGO AND GROUND HANDLING SERVICES PRIVATE LIMITED"
            )

            pdf.drawRightString(
                width - 38,
                terms_y - 35,
                "Authorized Signatory"
            )

            pdf.drawRightString(
                width - 38,
                terms_y - 47,
                "E. & O.E"
            )

        pdf.setFont(
            "Helvetica",
            7
        )

        pdf.drawCentredString(
            width / 2,
            32,
            "This is a Computer Generated Copy"
        )

        if not is_last_page:
            pdf.showPage()

    pdf.save()

    buffer.seek(0)

    safe_voucher = "".join(
        c if c.isalnum() or c in "-_" else "_"
        for c in voucher_no
    )

    filename = (
        f"Goods_Received_Note_{safe_voucher}.pdf"
        if safe_voucher
        else "Goods_Received_Note.pdf"
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


# =========================================================
# START APP
# =========================================================

if __name__ == "__main__":

    init_db()

    print()
    print("======================================")
    print("       INVENTORY DASHBOARD")
    print("======================================")
    print()
    print("Open this in your browser:")
    print()
    print("http://127.0.0.1:5000")
    print()
    print("======================================")

    # Set debug=False when deploying publicly.
    app.run(
        debug=True,
        use_reloader=False
    )
